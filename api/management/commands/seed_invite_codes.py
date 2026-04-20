from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from api.models import InviteCode


class Command(BaseCommand):
    help = "Seed InviteCode rows from an xlsx file (codes in column D, rows 5+)."

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to the codes xlsx file')

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path']).expanduser()
        if not xlsx_path.is_file():
            raise CommandError(f"File not found: {xlsx_path}")

        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise CommandError("openpyxl is required. pip install openpyxl") from e

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        raw_codes = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 4:
                continue
            # Data rows have a sequence number in column C and a code in column D;
            # rows without a sequence number are headers/metadata and are skipped.
            sequence = row[2]
            value = row[3]
            if sequence is None or value is None:
                continue
            code = str(value).strip()
            if not code or not code.isdigit() or len(code) < 6:
                continue
            raw_codes.append(code)

        if not raw_codes:
            raise CommandError("No codes found in column D of the xlsx")

        unique_codes = list(dict.fromkeys(raw_codes))
        existing = set(
            InviteCode.objects.filter(code__in=unique_codes).values_list('code', flat=True)
        )
        new_objs = [InviteCode(code=c) for c in unique_codes if c not in existing]

        InviteCode.objects.bulk_create(new_objs, ignore_conflicts=True)

        self.stdout.write(self.style.SUCCESS(
            f"Seeded {len(new_objs)} new codes, skipped {len(existing)} existing. "
            f"Total in DB: {InviteCode.objects.count()}"
        ))
